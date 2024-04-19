import numpy as np
import openpyxl as xl


def evolution_of_belief(bel0, pmatrix, n):
    bel = np.zeros((n + 1, 3))
    bel[0, 0:3] = bel0

    for i in range(n):
        bel[i + 1, :] = bel[i, :] @ pmatrix
    return bel


p = np.array([[1, 0, 0], [.4, .4, .2], [0, .6, .4]])
Bel0 = np.array([1, 0, 0])
Initial1 = evolution_of_belief(Bel0, p, 25)

Bel0 = np.array([0, 1, 0])
Initial2 = evolution_of_belief(Bel0, p, 25)

Bel0 = np.array([0, 0, 1])
Initial3 = evolution_of_belief(Bel0, p, 25)

wb = xl.load_workbook('SimulationBel.xlsx')
ws = wb.active
# save first result
for row in range(26):
    for col in range(3):
        a = ws.cell(row + 1, col + 1)
        a.value = Initial1[row, col]

# save 2nd result
for row in range(26):
    for col in range(3):
        a = ws.cell(row + 1, col + 5)
        a.value = Initial2[row, col]

# save 3rd result
for row in range(26):
    for col in range(3):
        a = ws.cell(row + 1, col + 9)
        a.value = Initial3[row, col]
wb.save("SimulationBel2.xlsx")
